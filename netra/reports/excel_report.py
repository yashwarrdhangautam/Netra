"""
netra/reports/excel_report.py
9-sheet openpyxl workbook for structured finding data:
  1. Summary          — risk scorecard and key metrics
  2. Findings         — full findings table, sortable
  3. Critical & High  — filtered high-priority view
  4. Assets           — all discovered assets
  5. Attack Chains    — multi-step attack paths
  6. MITRE Mapping    — findings mapped to ATT&CK
  7. Compliance       — HIPAA / PCI control status
  8. Remediation      — prioritised remediation roadmap
  9. Raw Data         — machine-readable findings dump
"""

import re
import json
from pathlib import Path
from datetime import datetime
from typing import List


# ── Colour palette ────────────────────────────────────────────────────
NAVY_HEX     = "1F3864"
BLUE_HEX     = "2E75B6"
CRITICAL_HEX = "FF0000"
HIGH_HEX     = "FF6600"
MEDIUM_HEX   = "FFC000"
LOW_HEX      = "92D050"
INFO_HEX     = "888888"
WHITE_HEX    = "FFFFFF"
LGREY_HEX    = "F2F2F2"
DGREY_HEX    = "CCCCCC"

SEV_FILLS = {
    "critical": CRITICAL_HEX,
    "high":     HIGH_HEX,
    "medium":   MEDIUM_HEX,
    "low":      LOW_HEX,
    "info":     INFO_HEX,
}

GRADE_FILLS = {"A": "00C864", "B": "2E75B6", "C": "FFC000", "D": "FF4444"}


def generate_excel_report(ctx: dict, reports_dir: Path) -> Path:
    """
    Generate a 9-sheet Excel workbook for the scan results.

    Args:
        ctx:         Report context dict.
        reports_dir: Directory to save the output file.

    Returns:
        Path to the generated .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        raise ImportError("openpyxl required. Run: pip3 install openpyxl --break-system-packages")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    findings  = ctx.get("findings", [])
    assets    = ctx.get("assets", [])
    chains    = ctx.get("chains", [])
    stats     = ctx.get("stats", {})
    score     = ctx.get("risk_score", 0)
    grade     = ctx.get("risk_grade", "?")

    _sheet_summary(wb, ctx, stats, score, grade)
    _sheet_findings(wb, findings, "Findings")
    _sheet_findings(wb, [f for f in findings
                         if f.get("severity") in ("critical", "high")],
                    "Critical & High")
    _sheet_assets(wb, assets)
    _sheet_chains(wb, chains, findings)
    _sheet_mitre(wb, findings)
    _sheet_compliance(wb, findings)
    _sheet_roadmap(wb, findings)
    _sheet_raw(wb, findings)

    filename = _report_filename(ctx, "xlsx")
    out_path = reports_dir / filename
    wb.save(str(out_path))
    return out_path


# ── Sheet builders ────────────────────────────────────────────────────

def _sheet_summary(wb, ctx: dict, stats: dict, score: int, grade: str) -> None:
    """Build the Summary sheet with KPIs and a bar chart."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Summary")

    # Title
    ws["B2"] = "NETRA Security Assessment — Summary"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True,
                         color=NAVY_HEX)

    # Meta info
    meta_rows = [
        ("Client",       ctx.get("client", "Confidential")),
        ("Engagement",   ctx.get("engagement", "Security Assessment")),
        ("Date",         ctx.get("date", datetime.now().strftime("%Y-%m-%d"))),
        ("Operator",     ctx.get("operator", "Security Team")),
        ("Scan ID",      ctx.get("scan_id", "N/A")),
    ]
    for i, (label, value) in enumerate(meta_rows, 4):
        ws.cell(i, 2, label).font = Font(bold=True)
        ws.cell(i, 3, str(value))

    # Risk grade
    ws["B10"] = "Risk Grade"
    ws["B10"].font = Font(bold=True)
    grade_cell = ws["C10"]
    grade_cell.value = f"{grade} ({score}/100)"
    grade_cell.font = Font(bold=True, size=14,
                            color=GRADE_FILLS.get(grade, NAVY_HEX))

    # Severity breakdown table
    ws["B12"] = "Severity"
    ws["C12"] = "Count"
    for cell in [ws["B12"], ws["C12"]]:
        _header_style(cell)

    sev_rows = [
        ("Critical", stats.get("critical", 0)),
        ("High",     stats.get("high", 0)),
        ("Medium",   stats.get("medium", 0)),
        ("Low",      stats.get("low", 0)),
        ("Info",     stats.get("info", 0)),
    ]
    for i, (sev, cnt) in enumerate(sev_rows, 13):
        ws.cell(i, 2, sev)
        ws.cell(i, 3, cnt)
        fill_hex = SEV_FILLS.get(sev.lower(), INFO_HEX)
        ws.cell(i, 2).fill = PatternFill("solid", fgColor=fill_hex)
        ws.cell(i, 2).font = Font(color=WHITE_HEX if sev in ("Critical", "High") else "000000")

    # Bar chart
    chart = BarChart()
    chart.title        = "Findings by Severity"
    chart.style        = 10
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.width        = 15
    chart.height       = 10

    data = Reference(ws, min_col=3, min_row=12, max_row=17)
    cats = Reference(ws, min_col=2, min_row=13, max_row=17)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E12")

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15


def _sheet_findings(wb, findings: list, sheet_name: str) -> None:
    """Build a sortable findings table sheet."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(sheet_name)

    headers = [
        "ID", "Title", "Severity", "CVSS", "CVE", "CWE",
        "Host", "URL", "Category", "OWASP Web", "MITRE",
        "HIPAA", "PCI", "Status", "Confidence", "Description"
    ]

    for j, h in enumerate(headers, 1):
        cell = ws.cell(1, j, h)
        _header_style(cell)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"

    for i, f in enumerate(findings, 2):
        row_data = [
            f.get("id"),
            f.get("title"),
            (f.get("severity") or "").capitalize(),
            f.get("cvss_score"),
            f.get("cve_id"),
            f.get("cwe_id"),
            f.get("host"),
            f.get("url"),
            f.get("category"),
            f.get("owasp_web"),
            f.get("mitre_technique"),
            f.get("hipaa_ref"),
            f.get("pci_ref"),
            f.get("status"),
            f.get("confidence"),
            (f.get("description") or "")[:200],
        ]
        for j, val in enumerate(row_data, 1):
            ws.cell(i, j, val or "")

        sev = (f.get("severity") or "info").lower()
        sev_fill = SEV_FILLS.get(sev, INFO_HEX)
        sev_cell = ws.cell(i, 3)
        sev_cell.fill = PatternFill("solid", fgColor=sev_fill)
        if sev in ("critical", "high"):
            sev_cell.font = Font(color=WHITE_HEX)

    _auto_width(ws)


def _sheet_assets(wb, assets: list) -> None:
    """Build the asset inventory sheet."""
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Assets")

    headers = ["Value", "Type", "Live", "Product", "Tech Stack", "Ports", "First Seen"]
    for j, h in enumerate(headers, 1):
        _header_style(ws.cell(1, j, h))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"

    for i, a in enumerate(assets, 2):
        try:
            ts_dict = json.loads(a.get("tech_stack") or "{}")
            ts_str  = ", ".join(f"{k}:{v}" for k, v in ts_dict.items())[:80]
        except Exception:
            ts_str  = str(a.get("tech_stack") or "")

        try:
            ports_list = json.loads(a.get("ports") or "[]")
            ports_str  = ", ".join(str(p) for p in ports_list[:10])
        except Exception:
            ports_str  = str(a.get("ports") or "")

        ws.cell(i, 1, a.get("value") or "")
        ws.cell(i, 2, a.get("asset_type") or "")
        ws.cell(i, 3, "Yes" if a.get("is_live") else "No")
        ws.cell(i, 4, a.get("product") or "")
        ws.cell(i, 5, ts_str)
        ws.cell(i, 6, ports_str)
        ws.cell(i, 7, str(a.get("first_seen") or "")[:16])

        if a.get("is_live"):
            ws.cell(i, 3).fill = PatternFill("solid", fgColor=LOW_HEX)

    _auto_width(ws)


def _sheet_chains(wb, chains: list, all_findings: list) -> None:
    """Build the attack chains sheet."""
    from openpyxl.styles import PatternFill, Font

    ws = wb.create_sheet("Attack Chains")
    id_map = {f["id"]: f for f in all_findings}

    headers = ["Chain #", "Combined CVSS", "MITRE Sequence", "Step Count",
               "Start Host", "Narrative", "Finding IDs"]
    for j, h in enumerate(headers, 1):
        _header_style(ws.cell(1, j, h))

    for i, c in enumerate(chains, 2):
        try:
            node_ids = json.loads(c.get("nodes", "[]")) if isinstance(c.get("nodes"), str) else c.get("nodes", [])
        except Exception:
            node_ids = []

        start_host = ""
        if node_ids:
            first_f    = id_map.get(node_ids[0])
            start_host = first_f.get("host", "") if first_f else ""

        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, c.get("combined_cvss"))
        ws.cell(i, 3, c.get("mitre_sequence", ""))
        ws.cell(i, 4, len(node_ids))
        ws.cell(i, 5, start_host)
        ws.cell(i, 6, (c.get("narrative") or "")[:300])
        ws.cell(i, 7, ", ".join(str(n) for n in node_ids))

        if c.get("combined_cvss") and c["combined_cvss"] >= 9:
            ws.cell(i, 2).fill = PatternFill("solid", fgColor=CRITICAL_HEX)
            ws.cell(i, 2).font = Font(color=WHITE_HEX, bold=True)

    _auto_width(ws)


def _sheet_mitre(wb, findings: list) -> None:
    """Build the MITRE ATT&CK mapping sheet."""
    ws = wb.create_sheet("MITRE Mapping")

    headers = ["MITRE Technique", "Count", "Findings", "Severity Distribution"]
    for j, h in enumerate(headers, 1):
        _header_style(ws.cell(1, j, h))

    # Group by technique
    technique_map: dict = {}
    for f in findings:
        t = f.get("mitre_technique", "").strip()
        if t:
            technique_map.setdefault(t, []).append(f)

    for i, (technique, flist) in enumerate(
        sorted(technique_map.items(), key=lambda x: len(x[1]), reverse=True), 2
    ):
        sev_dist = ", ".join(
            f"{s}:{sum(1 for f in flist if f.get('severity') == s)}"
            for s in ["critical", "high", "medium", "low"]
            if any(f.get("severity") == s for f in flist)
        )
        ws.cell(i, 1, technique)
        ws.cell(i, 2, len(flist))
        ws.cell(i, 3, ", ".join(str(f.get("id")) for f in flist[:5]))
        ws.cell(i, 4, sev_dist)

    _auto_width(ws)


def _sheet_compliance(wb, findings: list) -> None:
    """Build the compliance gap analysis sheet."""
    from openpyxl.styles import PatternFill, Font

    ws = wb.create_sheet("Compliance")

    headers = ["Framework", "Control", "Status", "Finding Count", "Notes"]
    for j, h in enumerate(headers, 1):
        _header_style(ws.cell(1, j, h))

    row = 2
    for framework in ("HIPAA", "PCI"):
        try:
            from netra.ai_brain.analyzer import compute_compliance_gaps
            gaps = compute_compliance_gaps(findings, framework)
        except Exception:
            gaps = []

        for gap in gaps:
            ws.cell(row, 1, gap["framework"])
            ws.cell(row, 2, gap["control"])
            ws.cell(row, 3, gap["status"].upper())
            ws.cell(row, 4, len(gap.get("finding_ids", [])))
            ws.cell(row, 5, gap.get("notes", "")[:100])

            status_cell = ws.cell(row, 3)
            if gap["status"] == "fail":
                status_cell.fill = PatternFill("solid", fgColor=CRITICAL_HEX)
                status_cell.font = Font(color=WHITE_HEX, bold=True)
            elif gap["status"] == "pass":
                status_cell.fill = PatternFill("solid", fgColor=LOW_HEX)
            elif gap["status"] == "warn":
                status_cell.fill = PatternFill("solid", fgColor=MEDIUM_HEX)

            row += 1

    _auto_width(ws)


def _sheet_roadmap(wb, findings: list) -> None:
    """Build the remediation roadmap sheet."""
    from openpyxl.styles import PatternFill, Font

    ws = wb.create_sheet("Remediation")

    headers = ["Priority", "Severity", "Count", "Effort", "SLA", "Finding Titles"]
    for j, h in enumerate(headers, 1):
        _header_style(ws.cell(1, j, h))

    try:
        from netra.ai_brain.analyzer import build_remediation_roadmap
        roadmap = build_remediation_roadmap(findings)
    except Exception:
        roadmap = []

    sla_map = {
        "critical": "< 24 hours",
        "high":     "< 72 hours",
        "medium":   "1-2 sprints",
        "low":      "Next release",
    }

    for i, item in enumerate(roadmap, 2):
        ws.cell(i, 1, item["priority"])
        ws.cell(i, 2, item["severity"].capitalize())
        ws.cell(i, 3, item["count"])
        ws.cell(i, 4, item["effort"])
        ws.cell(i, 5, sla_map.get(item["severity"], "TBD"))
        ws.cell(i, 6, "; ".join(item["top_titles"][:3])[:150])

        sev_cell = ws.cell(i, 2)
        sev_cell.fill = PatternFill("solid", fgColor=SEV_FILLS.get(item["severity"], INFO_HEX))
        if item["severity"] in ("critical", "high"):
            sev_cell.font = Font(color=WHITE_HEX, bold=True)

    _auto_width(ws)


def _sheet_raw(wb, findings: list) -> None:
    """Build the raw data sheet with all finding fields."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Raw Data")

    all_keys = [
        "id", "scan_id", "title", "severity", "cvss_score", "cvss_vector",
        "cve_id", "cwe_id", "category", "owasp_web", "owasp_api", "owasp_llm",
        "mitre_technique", "hipaa_ref", "pci_ref",
        "host", "url", "path", "parameter", "product",
        "description", "evidence", "poc_command",
        "impact", "remediation", "ai_narrative",
        "status", "confidence", "operator",
        "found_at", "updated_at",
    ]

    for j, key in enumerate(all_keys, 1):
        _header_style(ws.cell(1, j, key))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_keys))}1"
    ws.freeze_panes    = "A2"

    for i, f in enumerate(findings, 2):
        for j, key in enumerate(all_keys, 1):
            val = f.get(key, "")
            if val is None:
                val = ""
            ws.cell(i, j, str(val)[:500] if isinstance(val, str) else val)

    _auto_width(ws)


# ── Style helpers ─────────────────────────────────────────────────────

def _header_style(cell) -> None:
    """Apply standard header styling to a cell."""
    from openpyxl.styles import PatternFill, Font, Alignment

    cell.fill      = PatternFill("solid", fgColor=NAVY_HEX)
    cell.font      = Font(name="Calibri", bold=True, color=WHITE_HEX, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _auto_width(ws, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _report_filename(ctx: dict, ext: str) -> str:
    """Generate standardised report filename."""
    scan_id = ctx.get("scan_id", "unknown")
    date    = ctx.get("date", datetime.now().strftime("%Y%m%d")).replace("-", "")
    target  = re.sub(r"[^a-zA-Z0-9]", "_", ctx.get("engagement", scan_id))[:20]
    return f"NETRA_vapt_{target}_{date}.{ext}"
