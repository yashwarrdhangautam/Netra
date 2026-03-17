"""
recon/extractor.py
Asset extraction from all input sources.
Reads Excel (WestCX format), TXT files, paste, or editor.
Separates IPs from domains, tags by product, applies scope filter.
"""

import re
import sys
import tempfile
import subprocess
from pathlib import Path
from typing import Tuple

from netra.core.config import CONFIG, filter_targets
from netra.core.utils  import status, banner, is_ip, is_domain, is_url, deduplicate, C


class AssetExtractor:
    """
    Extracts and categorizes targets from multiple input sources.

    Result:
        .domains  — list of domain names
        .ips      — list of IP addresses / CIDR ranges
        .urls     — list of full URLs
        .by_product — dict: {product_name: [targets]}
        .all_targets — flat combined list
    """

    def __init__(self):
        self.domains:    list = []
        self.ips:        list = []
        self.urls:       list = []
        self.by_product: dict = {}
        self.raw:        list = []

    def from_excel(self, path: str) -> "AssetExtractor":
        """
        Read targets from Excel file.
        Expected: sheet per product, column with domains/IPs.
        Auto-detects column headers: domain, ip, target, host, url, asset.
        """
        try:
            import openpyxl
        except ImportError:
            status("openpyxl not installed — run: pip3 install openpyxl", "error")
            return self

        status(f"Reading Excel: {Path(path).name}", "run")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws      = wb[sheet_name]
            product = sheet_name.strip()
            if product.lower() in ("sheet1", "sheet2", "overview", "summary"):
                product = ""

            col_idx = None
            headers = []

            for row in ws.iter_rows(values_only=True):
                if col_idx is None:
                    # Find the right column on first row
                    headers = [str(c).lower().strip() if c else "" for c in row]
                    for i, h in enumerate(headers):
                        if any(k in h for k in
                               ("domain", "ip", "target", "host", "url", "asset", "address")):
                            col_idx = i
                            break
                    if col_idx is None and row:
                        col_idx = 0   # fallback to first column
                    continue

                if row and col_idx < len(row):
                    val = row[col_idx]
                    if val:
                        target = str(val).strip()
                        if target and not target.lower().startswith(("http", "example", "n/a", "none")):
                            self.raw.append(target)
                            if product:
                                self.by_product.setdefault(product, []).append(target)

        wb.close()
        self._classify()
        status(f"Excel: {len(self.domains)} domains, {len(self.ips)} IPs "
               f"from {len(self.by_product)} products", "ok")
        return self

    def from_file(self, path: str, product: str = "") -> "AssetExtractor":
        """Read from a plain text file, one target per line."""
        status(f"Reading file: {Path(path).name}", "run")
        try:
            lines = Path(path).read_text().splitlines()
        except Exception as e:
            status(f"Cannot read {path}: {e}", "error")
            return self

        for line in lines:
            line = line.strip()
            if line and not line.startswith("#"):
                self.raw.append(line)
                if product:
                    self.by_product.setdefault(product, []).append(line)

        self._classify()
        status(f"File: {len(self.domains)} domains, {len(self.ips)} IPs", "ok")
        return self

    def from_paste(self, text: str, product: str = "") -> "AssetExtractor":
        """Parse targets pasted directly as a string."""
        # Split on common separators
        for sep in (",", ";", "\n", " "):
            if sep in text:
                items = [t.strip() for t in text.split(sep)]
                break
        else:
            items = [text.strip()]

        for item in items:
            item = item.strip()
            if item:
                self.raw.append(item)
                if product:
                    self.by_product.setdefault(product, []).append(item)

        self._classify()
        return self

    def from_editor(self) -> "AssetExtractor":
        """Open $EDITOR for user to type/paste targets. Returns when editor closes."""
        instructions = (
            "# Enter one target per line (domain, IP, or URL)\n"
            "# Lines starting with # are ignored\n"
            "# Product tagging: [ProductName] on its own line groups targets below it\n"
            "# Example:\n"
            "# [Mosaicx]\n"
            "# mosaicx.com\n"
            "# 192.168.1.0/24\n"
            "#\n"
        )
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".txt", prefix="netra_targets_",
            delete=False
        ) as tmp:
            tmp.write(instructions)
            tmp_path = tmp.name

        editor = (
            subprocess.run(["git", "config", "--global", "core.editor"],
                           capture_output=True, text=True).stdout.strip()
            or "nano"
        )

        subprocess.run([editor, tmp_path])

        current_product = ""
        for line in Path(tmp_path).read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # Product tag: [ProductName]
            m = re.match(r"^\[(.+)\]$", line)
            if m:
                current_product = m.group(1).strip()
                continue
            self.raw.append(line)
            if current_product:
                self.by_product.setdefault(current_product, []).append(line)

        Path(tmp_path).unlink(missing_ok=True)
        self._classify()
        status(f"Editor: {len(self.domains)} domains, {len(self.ips)} IPs", "ok")
        return self

    def _classify(self) -> None:
        """Classify raw targets into domains / IPs / URLs. Apply scope filter."""
        excluded = set(t.lower() for t in
                       CONFIG.get("exclude_domains", "").replace(";", ",").split(",")
                       if t.strip())
        excluded |= set(t.lower() for t in
                        CONFIG.get("exclude_ips", "").replace(";", ",").split(",")
                        if t.strip())

        include_private = str(CONFIG.get("include_private", "false")).lower() == "true"

        all_raw = deduplicate(self.raw)

        for t in all_raw:
            t_clean = t.strip().lower()

            # Skip excluded
            if any(ex in t_clean or t_clean == ex for ex in excluded if ex):
                continue

            if is_url(t):
                self.urls.append(t)
            elif is_ip(t) or "/" in t:  # IP or CIDR
                from netra.core.utils import is_private_ip
                if not include_private and is_private_ip(t.split("/")[0]):
                    continue
                self.ips.append(t)
            elif is_domain(t):
                self.domains.append(t)
            else:
                # Try treating as domain anyway
                if "." in t:
                    self.domains.append(t)

        self.domains = deduplicate(self.domains)
        self.ips     = deduplicate(self.ips)
        self.urls    = deduplicate(self.urls)

    @property
    def all_targets(self) -> list:
        """Return all targets (domains + IPs + URLs) as a flat list."""
        return self.domains + self.ips + self.urls

    def save(self, workdir: str) -> dict:
        """Write categorized targets to workdir files."""
        workdir = Path(workdir)
        paths   = {}

        if self.domains:
            p = workdir / "recon" / "domains.txt"
            p.write_text("\n".join(self.domains) + "\n")
            paths["domains"] = str(p)

        if self.ips:
            p = workdir / "recon" / "ips.txt"
            p.write_text("\n".join(self.ips) + "\n")
            paths["ips"] = str(p)

        if self.urls:
            p = workdir / "recon" / "urls.txt"
            p.write_text("\n".join(self.urls) + "\n")
            paths["urls"] = str(p)

        if self.by_product:
            product_dir = workdir / "recon" / "by_product"
            product_dir.mkdir(exist_ok=True)
            for product, targets in self.by_product.items():
                slug = re.sub(r"[^a-zA-Z0-9_-]", "_", product)[:40]
                p    = product_dir / f"{slug}.txt"
                p.write_text("\n".join(deduplicate(targets)) + "\n")
            paths["by_product"] = str(product_dir)

        return paths

    def summary(self) -> str:
        """Return a human-readable summary of parsed target counts."""
        lines = [
            f"  Domains:  {len(self.domains)}",
            f"  IPs:      {len(self.ips)}",
            f"  URLs:     {len(self.urls)}",
            f"  Products: {len(self.by_product)} ({', '.join(list(self.by_product.keys())[:5])})",
            f"  Total:    {len(self.all_targets)}",
        ]
        return "\n".join(lines)


def interactive_input_menu() -> AssetExtractor:
    """
    Interactive menu for selecting input method.
    Shown on startup when no targets are provided.
    """
    banner("TARGET INPUT", "How would you like to provide targets?")

    options = [
        ("1", "Excel file (.xlsx)",           "from_excel"),
        ("2", "Text file (.txt)",             "from_file"),
        ("3", "Paste targets directly",       "from_paste"),
        ("4", "Open editor (nano/vim)",       "from_editor"),
    ]

    for num, label, _ in options:
        print(f"  [{num}] {label}")
    print()

    choice = input("  Select [1-4]: ").strip()
    extractor = AssetExtractor()

    if choice == "1":
        path = input("  Excel file path: ").strip().strip("'\"")
        extractor.from_excel(path)

    elif choice == "2":
        path    = input("  Text file path: ").strip().strip("'\"")
        product = input("  Product name (optional): ").strip()
        extractor.from_file(path, product)

    elif choice == "3":
        print("  Paste targets (domains, IPs, URLs). Press Enter twice when done:")
        lines = []
        while True:
            line = input()
            if not line and lines:
                break
            lines.append(line)
        product = input("  Product name (optional): ").strip()
        extractor.from_paste("\n".join(lines), product)

    elif choice == "4":
        extractor.from_editor()

    else:
        print(f"  {C.RED}Invalid choice{C.RESET}")
        return interactive_input_menu()

    print()
    print(extractor.summary())

    if not extractor.all_targets:
        status("No targets found. Try again.", "error")
        return interactive_input_menu()

    return extractor
